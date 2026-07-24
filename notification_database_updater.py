from __future__ import annotations

import argparse
import csv
import logging
import os
import tempfile
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# パス設定
# ============================================================

# このファイルをプロジェクトルート直下へ配置する
PROJECT_ROOT = Path(__file__).resolve().parent

STATIC_DIR = PROJECT_ROOT / "static"
RESULT_DIR = STATIC_DIR / "result"

# 既存のExcelファイルをresultディレクトリへ配置する
DATABASE_XLSX = RESULT_DIR / "notification_database.xlsx"


# ============================================================
# CSVと更新対象シートの対応
# ============================================================

# notificationシートは今回一切更新しない。
SOURCE_SETTINGS: dict[str, dict[str, Any]] = {
    "detections.csv": {
        "sheet_name": "realtime_sheet",
        "columns": [
            "timestamp",
            "device_id",
            "animal_type",
            "confidence",
            "action_triggered",
            "stay_duration",
        ],
    },
    "daily_analysis.csv": {
        "sheet_name": "daily_sheet",
        "columns": [
            "date",
            "animal",
            "total_count",
            "average_stay_time",
            "alert_level",
        ],
    },
    "weekly_analysis.csv": {
        "sheet_name": "weekly_sheet",
        "columns": [
            "week",
            "animal",
            "peak_hour",
            "peak_day",
            "familiarity_score",
            "trap_score",
            "level",
        ],
    },
    "monthly_analysis.csv": {
        "sheet_name": "monthly_sheet",
        "columns": [
            "month",
            "device_id",
            "total_count",
            "monkey_count",
            "boar_count",
            "trap_score",
            "rank",
        ],
    },
    "yearly_analysis.csv": {
        "sheet_name": "yearly_sheet",
        "columns": [
            "year",
            "device_id",
            "total_count",
            "monkey_count",
            "boar_count",
            "trap_score",
            "rank",
        ],
    },
}


# 今回触らない計算・通知用シート
UNTOUCHED_NOTIFICATION_SHEETS = [
    "realtime_notification",
    "daily_notification",
    "weekly_notification",
    "monthly_notification",
    "yearly_notification",
]


# ============================================================
# ログ設定
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

LOGGER = logging.getLogger(__name__)


# ============================================================
# 入力ファイル・Excel構成の検証
# ============================================================

def validate_required_files() -> None:
    """
    static配下の入力CSVと、
    static/result/notification_database.xlsxの存在を確認する。

    notification_database.xlsxはコードから新規作成しない。
    """

    if not STATIC_DIR.exists():
        raise FileNotFoundError(
            f"staticディレクトリがありません: {STATIC_DIR}"
        )

    RESULT_DIR.mkdir(parents=True, exist_ok=True)

    if not DATABASE_XLSX.exists():
        raise FileNotFoundError(
            "既存のnotification_database.xlsxがありません。\n"
            f"次の場所へ配置してください: {DATABASE_XLSX}"
        )

    if not DATABASE_XLSX.is_file():
        raise ValueError(
            f"Excelのパスがファイルではありません: {DATABASE_XLSX}"
        )

    missing_csv_files = [
        filename
        for filename in SOURCE_SETTINGS
        if not (STATIC_DIR / filename).is_file()
    ]

    if missing_csv_files:
        missing_text = "\n".join(
            f"- {STATIC_DIR / filename}"
            for filename in missing_csv_files
        )
        raise FileNotFoundError(
            f"次の入力CSVがありません:\n{missing_text}"
        )


def validate_workbook_sheets(workbook) -> None:
    """
    更新対象の5シートと、触らない通知用5シートが
    既存Excel内に存在することを確認する。

    不足シートは自動作成しない。
    """

    required_sheets = [
        setting["sheet_name"]
        for setting in SOURCE_SETTINGS.values()
    ] + UNTOUCHED_NOTIFICATION_SHEETS

    missing_sheets = [
        sheet_name
        for sheet_name in required_sheets
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        missing_text = "\n".join(
            f"- {sheet_name}"
            for sheet_name in missing_sheets
        )
        raise ValueError(
            "notification_database.xlsxに必要なシートがありません。\n"
            f"{missing_text}"
        )


# ============================================================
# CSV読み込み
# ============================================================

def read_csv_file(
    csv_path: Path,
    expected_columns: list[str],
) -> list[dict[str, str]]:
    """
    CSVを読み込み、ヘッダー名と項目順を厳密に検証する。

    UTF-8およびUTF-8 BOM付きCSVに対応する。
    """

    with csv_path.open(
        mode="r",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        reader = csv.DictReader(file)

        if reader.fieldnames is None:
            raise ValueError(
                f"CSVにヘッダーがありません: {csv_path}"
            )

        actual_columns = [
            column.strip()
            for column in reader.fieldnames
        ]

        if actual_columns != expected_columns:
            missing_columns = [
                column
                for column in expected_columns
                if column not in actual_columns
            ]
            extra_columns = [
                column
                for column in actual_columns
                if column not in expected_columns
            ]

            raise ValueError(
                "CSV項目が仕様と一致していません。\n"
                f"対象ファイル: {csv_path}\n"
                f"期待する項目: {expected_columns}\n"
                f"実際の項目: {actual_columns}\n"
                f"不足項目: {missing_columns}\n"
                f"余分な項目: {extra_columns}"
            )

        rows: list[dict[str, str]] = []

        for line_number, raw_row in enumerate(reader, start=2):
            normalized_row = {
                key.strip(): (
                    value.strip()
                    if isinstance(value, str)
                    else ""
                )
                for key, value in raw_row.items()
                if key is not None
            }

            # 完全な空行は無視する
            if all(
                normalized_row.get(column, "") == ""
                for column in expected_columns
            ):
                continue

            missing_values = [
                column
                for column in expected_columns
                if column not in normalized_row
            ]

            if missing_values:
                raise ValueError(
                    f"{csv_path.name}の{line_number}行目に"
                    f"不足項目があります: {missing_values}"
                )

            rows.append({
                column: normalized_row.get(column, "")
                for column in expected_columns
            })

    LOGGER.info(
        "%sから%d件読み込みました。",
        csv_path.name,
        len(rows),
    )

    return rows


# ============================================================
# Excel用の型変換
# ============================================================

def convert_excel_value(
    column_name: str,
    value: str,
) -> Any:
    """
    CSV文字列を、Excelで扱いやすい数値型へ変換する。
    """

    if value == "":
        return None

    integer_columns = {
        "stay_duration",
        "total_count",
        "monkey_count",
        "boar_count",
        "year",
    }

    float_columns = {
        "confidence",
        "average_stay_time",
        "familiarity_score",
        "trap_score",
    }

    if column_name in integer_columns:
        try:
            return int(float(value))
        except ValueError as error:
            raise ValueError(
                f"{column_name}には整数が必要です。入力値={value}"
            ) from error

    if column_name in float_columns:
        try:
            return float(value)
        except ValueError as error:
            raise ValueError(
                f"{column_name}には数値が必要です。入力値={value}"
            ) from error

    return value


# ============================================================
# rawシート更新
# ============================================================

def replace_raw_sheet_data(
    worksheet: Worksheet,
    columns: list[str],
    rows: list[dict[str, str]],
) -> None:
    """
    realtime_sheet、daily_sheet、weekly_sheet、
    monthly_sheet、yearly_sheetだけを最新データへ置き換える。

    notificationシートはこの関数から触らない。
    """

    # 古いrawデータを削除する。
    # シート自体は削除しない。
    if worksheet.max_row > 0:
        worksheet.delete_rows(
            idx=1,
            amount=worksheet.max_row,
        )

    # CSVと同じ項目名・項目順を設定する
    worksheet.append(columns)

    for row in rows:
        worksheet.append([
            convert_excel_value(
                column_name=column,
                value=row.get(column, ""),
            )
            for column in columns
        ])

    format_raw_sheet(
        worksheet=worksheet,
        columns=columns,
    )


def format_raw_sheet(
    worksheet: Worksheet,
    columns: list[str],
) -> None:
    """
    更新したrawシートだけに最低限の表示形式を設定する。
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_positions = {
        column: index + 1
        for index, column in enumerate(columns)
    }

    if "confidence" in column_positions:
        column_index = column_positions["confidence"]
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = "0.0000"

    for numeric_column in [
        "average_stay_time",
        "familiarity_score",
        "trap_score",
    ]:
        if numeric_column not in column_positions:
            continue

        column_index = column_positions[numeric_column]
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(
                row=row_number,
                column=column_index,
            ).number_format = "0.00"

    # 過度に広くならない範囲で列幅を調整する
    for column_cells in worksheet.columns:
        maximum_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(cell.value)),
            )

        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(
            max(maximum_length + 2, 12),
            35,
        )


# ============================================================
# Excelの安全な保存
# ============================================================

def save_workbook_atomically(workbook) -> None:
    """
    一時Excelの保存に成功してから、
    static/result/notification_database.xlsxを置き換える。

    保存途中の破損を避けるため、直接上書きしない。
    """

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix="notification_database_",
        suffix=".xlsx",
        dir=RESULT_DIR,
    )

    os.close(file_descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)

        os.replace(
            temporary_path,
            DATABASE_XLSX,
        )

    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()

        raise


# ============================================================
# 全体更新処理
# ============================================================

def update_notification_database() -> None:
    """
    5つのCSVを読み込み、既存Excelのrawシートだけを更新する。

    更新対象:
    - realtime_sheet
    - daily_sheet
    - weekly_sheet
    - monthly_sheet
    - yearly_sheet

    更新しない:
    - realtime_notification
    - daily_notification
    - weekly_notification
    - monthly_notification
    - yearly_notification

    notification_database.csvは作成・使用しない。
    """

    validate_required_files()

    # 先に全CSVを検証する。
    # 1ファイルでも不正ならExcelを更新しない。
    source_data: dict[str, list[dict[str, str]]] = {}

    for source_file, setting in SOURCE_SETTINGS.items():
        source_data[source_file] = read_csv_file(
            csv_path=STATIC_DIR / source_file,
            expected_columns=setting["columns"],
        )

    workbook = load_workbook(DATABASE_XLSX)

    try:
        validate_workbook_sheets(workbook)

        for source_file, setting in SOURCE_SETTINGS.items():
            sheet_name = setting["sheet_name"]
            worksheet = workbook[sheet_name]

            replace_raw_sheet_data(
                worksheet=worksheet,
                columns=setting["columns"],
                rows=source_data[source_file],
            )

            LOGGER.info(
                "%sを%sへ反映しました。",
                source_file,
                sheet_name,
            )

        save_workbook_atomically(workbook)

        LOGGER.info(
            "notification_database.xlsxを更新しました: %s",
            DATABASE_XLSX,
        )

        LOGGER.info(
            "notificationシート5枚は変更していません。"
        )

    finally:
        workbook.close()


# ============================================================
# CSV更新監視
# ============================================================

def get_input_file_states() -> dict[str, int]:
    """
    入力CSVの最終更新時刻を取得する。
    """

    return {
        filename: (
            (STATIC_DIR / filename).stat().st_mtime_ns
            if (STATIC_DIR / filename).exists()
            else 0
        )
        for filename in SOURCE_SETTINGS
    }


def watch_input_csv_files(
    interval_seconds: float,
) -> None:
    """
    CSVを定期監視し、更新があった場合に
    notification_database.xlsxのrawシートだけを更新する。
    """

    previous_states = get_input_file_states()

    LOGGER.info(
        "CSV監視を開始しました。監視間隔: %.1f秒",
        interval_seconds,
    )
    LOGGER.info(
        "終了する場合はCtrl+Cを押してください。"
    )

    try:
        while True:
            time.sleep(interval_seconds)

            current_states = get_input_file_states()

            if current_states == previous_states:
                continue

            changed_files = [
                filename
                for filename in SOURCE_SETTINGS
                if current_states.get(filename)
                != previous_states.get(filename)
            ]

            LOGGER.info(
                "更新されたCSVを検知しました: %s",
                ", ".join(changed_files),
            )

            # 別担当の処理がCSVを書き込み中である可能性を考慮する
            time.sleep(0.5)

            try:
                update_notification_database()
                previous_states = get_input_file_states()

            except Exception:
                LOGGER.exception(
                    "CSV更新後のExcel反映に失敗しました。"
                )

    except KeyboardInterrupt:
        LOGGER.info(
            "CSV監視を終了しました。"
        )


# ============================================================
# 起動処理
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "5つのCSVを読み込み、既存の"
            "static/result/notification_database.xlsxの"
            "rawシートだけを更新します。"
        )
    )

    parser.add_argument(
        "--watch",
        action="store_true",
        help="CSVを監視し、変更時にExcelを自動更新します。",
    )

    parser.add_argument(
        "--interval",
        type=float,
        default=2.0,
        help="監視間隔の秒数。既定値は2秒です。",
    )

    args = parser.parse_args()

    if args.interval < 0.5:
        raise ValueError(
            "監視間隔は0.5秒以上にしてください。"
        )

    # 起動時に1回更新する
    update_notification_database()

    if args.watch:
        watch_input_csv_files(
            interval_seconds=args.interval,
        )


if __name__ == "__main__":
    try:
        main()

    except PermissionError as error:
        LOGGER.error(
            "notification_database.xlsxへのアクセスが拒否されました。"
            "ExcelやLibreOfficeで開いている場合は閉じてください。"
        )
        raise SystemExit(1) from error

    except (FileNotFoundError, ValueError) as error:
        LOGGER.error("%s", error)
        raise SystemExit(1) from error

    except Exception as error:
        LOGGER.exception(
            "更新処理に失敗しました。"
        )
        raise SystemExit(1) from error