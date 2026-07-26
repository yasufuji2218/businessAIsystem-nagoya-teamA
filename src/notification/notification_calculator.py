from __future__ import annotations

import logging
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger(__name__)


# ============================================================
# シート名
# ============================================================

RAW_SHEETS = {
    "realtime": "realtime_sheet",
    "daily": "daily_sheet",
    "weekly": "weekly_sheet",
    "monthly": "monthly_sheet",
    "yearly": "yearly_sheet",
}

NOTIFICATION_SHEETS = {
    "realtime": "realtime_notification",
    "daily": "daily_notification",
    "weekly": "weekly_notification",
    "monthly": "monthly_notification",
    "yearly": "yearly_notification",
}


# ============================================================
# rawシートの必須項目
# ============================================================

RAW_REQUIRED_COLUMNS = {
    "realtime_sheet": [
        "timestamp",
        "device_id",
        "animal_type",
        "confidence",
        "action_triggered",
        "stay_duration",
    ],
    "daily_sheet": [
        "date",
        "animal",
        "total_count",
        "average_stay_time",
        "alert_level",
    ],
    "weekly_sheet": [
        "week",
        "animal",
        "peak_hour",
        "peak_day",
        "familiarity_score",
        "trap_score",
        "level",
    ],
    "monthly_sheet": [
        "month",
        "device_id",
        "total_count",
        "monkey_count",
        "boar_count",
        "trap_score",
        "rank",
    ],
    "yearly_sheet": [
        "year",
        "device_id",
        "total_count",
        "monkey_count",
        "boar_count",
        "trap_score",
        "rank",
    ],
}


# ============================================================
# notificationシートの出力項目
# ============================================================

NOTIFICATION_COLUMNS = {
    "realtime_notification": [
        "timestamp",
        "device_id",
        "animal_type",
        "confidence",
        "action_triggered",
        "stay_duration",
        "notification_status",
    ],
    "daily_notification": [
        "date",
        "total_detection_count",
        "boar_count",
        "monkey_count",
        "average_stay_time",
        "night_alert_level",
        "notification_status",
    ],
    "weekly_notification": [
        "week",
        "animal",
        "peak_hour",
        "peak_day",
        "familiarity_score",
        "trap_score",
        "level",
        "notification_status",
    ],
    "monthly_notification": [
        "month",
        "device_id",
        "monthly_total_count",
        "boar_ratio",
        "monkey_ratio",
        "comparison_previous_month",
        "monthly_peak_hour",
        "action_effectiveness",
        "trap_score",
        "notification_status",
    ],
    "yearly_notification": [
        "year",
        "device_id",
        "total_count",
        "monkey_count",
        "boar_count",
        "trap_score",
        "notification_status",
    ],
}


# ============================================================
# 通知状態
# ============================================================

PENDING = "PENDING"

VALID_NOTIFICATION_STATUSES = {
    "PENDING",
    "SUCCESS",
    "FAILED",
    "SKIPPED",
}


# ============================================================
# 警戒レベル
# ============================================================

ALERT_LEVEL_PRIORITY = {
    "LOW": 1,
    "MEDIUM": 2,
    "HIGH": 3,
}


# ============================================================
# 例外
# ============================================================

class NotificationCalculationError(RuntimeError):
    """通知用データの計算に失敗した場合の例外。"""


# ============================================================
# 基本変換
# ============================================================

def normalize_text(value: object) -> str:
    """セル値を前後空白のない文字列へ変換する。"""
    if value is None:
        return ""

    return str(value).strip()


def to_int(
    value: object,
    default: int = 0,
) -> int:
    """セル値を整数へ変換する。"""
    if value is None or value == "":
        return default

    if isinstance(value, bool):
        return int(value)

    try:
        return int(float(value))

    except (TypeError, ValueError) as error:
        raise NotificationCalculationError(
            f"整数へ変換できません。値={value!r}"
        ) from error


def to_float(
    value: object,
    default: float = 0.0,
) -> float:
    """セル値を浮動小数点数へ変換する。"""
    if value is None or value == "":
        return default

    try:
        return float(value)

    except (TypeError, ValueError) as error:
        raise NotificationCalculationError(
            f"数値へ変換できません。値={value!r}"
        ) from error


def normalize_date_value(value: object) -> str:
    """日付をYYYY-MM-DD形式へ変換する。"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = normalize_text(value)

    if not text:
        return ""

    # 日付時刻文字列の場合も先頭の日付部分だけ取得する。
    try:
        return datetime.fromisoformat(text).strftime("%Y-%m-%d")

    except ValueError:
        return text[:10]


def normalize_month_value(value: object) -> str:
    """年月をYYYY-MM形式へ変換する。"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")

    if isinstance(value, date):
        return value.strftime("%Y-%m")

    text = normalize_text(value)

    if not text:
        return ""

    try:
        return datetime.fromisoformat(text).strftime("%Y-%m")

    except ValueError:
        return text[:7]


def normalize_year_value(value: object) -> int:
    """年を整数へ変換する。"""
    if isinstance(value, datetime):
        return value.year

    if isinstance(value, date):
        return value.year

    return to_int(value)


def parse_timestamp(value: object) -> datetime | None:
    """timestampをdatetimeへ変換する。"""
    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time(),
        )

    text = normalize_text(value)

    if not text:
        return None

    try:
        return datetime.fromisoformat(text)

    except ValueError:
        pass

    supported_formats = [
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
    ]

    for date_format in supported_formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            )

        except ValueError:
            continue

    LOGGER.warning(
        "timestampを解析できませんでした: %r",
        value,
    )

    return None


# ============================================================
# Excel読込
# ============================================================

def read_worksheet_rows(
    worksheet: Worksheet,
    required_columns: list[str],
) -> list[dict[str, object]]:
    """
    Excelシートを辞書のリストとして読み込む。

    1行目をヘッダーとして使用する。
    """
    header_values = [
        normalize_text(cell.value)
        for cell in worksheet[1]
    ]

    actual_columns = [
        column
        for column in header_values
        if column
    ]

    if actual_columns != required_columns:
        missing_columns = [
            column
            for column in required_columns
            if column not in actual_columns
        ]

        extra_columns = [
            column
            for column in actual_columns
            if column not in required_columns
        ]

        raise NotificationCalculationError(
            "シートの項目が仕様と一致していません。\n"
            f"シート名: {worksheet.title}\n"
            f"期待する項目: {required_columns}\n"
            f"実際の項目: {actual_columns}\n"
            f"不足項目: {missing_columns}\n"
            f"余分な項目: {extra_columns}"
        )

    rows: list[dict[str, object]] = []

    for row_values in worksheet.iter_rows(
        min_row=2,
        max_col=len(required_columns),
        values_only=True,
    ):
        if all(value is None or value == "" for value in row_values):
            continue

        row = {
            column: row_values[index]
            for index, column in enumerate(required_columns)
        }

        rows.append(row)

    return rows


def validate_required_sheets(
    workbook: Any,
) -> None:
    """必要なrawシートとnotificationシートの存在を確認する。"""
    required_sheet_names = (
        list(RAW_SHEETS.values())
        + list(NOTIFICATION_SHEETS.values())
    )

    missing_sheets = [
        sheet_name
        for sheet_name in required_sheet_names
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise NotificationCalculationError(
            "notification_database.xlsxに"
            "必要なシートがありません。\n"
            + "\n".join(
                f"- {sheet_name}"
                for sheet_name in missing_sheets
            )
        )


# ============================================================
# 重複除去
# ============================================================

def keep_latest_rows(
    rows: list[dict[str, object]],
    key_columns: tuple[str, ...],
) -> list[dict[str, object]]:
    """
    指定キーが同一の行について、シート内で後ろにある行を残す。

    backend.batchが同じ期間を複数回追記した場合の
    重複集計を防止する。
    """
    latest_rows: dict[
        tuple[object, ...],
        dict[str, object],
    ] = {}

    for row in rows:
        key = tuple(
            normalize_text(row.get(column))
            for column in key_columns
        )

        latest_rows[key] = row

    return list(latest_rows.values())


def remove_exact_duplicate_rows(
    rows: list[dict[str, object]],
    columns: tuple[str, ...],
) -> list[dict[str, object]]:
    """すべての指定項目が同一の完全重複行だけを削除する。"""
    unique_rows: list[dict[str, object]] = []
    seen: set[tuple[str, ...]] = set()

    for row in rows:
        signature = tuple(
            normalize_text(row.get(column))
            for column in columns
        )

        if signature in seen:
            continue

        seen.add(signature)
        unique_rows.append(row)

    return unique_rows


# ============================================================
# 既存通知状態の維持
# ============================================================

def load_existing_statuses(
    worksheet: Worksheet,
    columns: list[str],
) -> dict[tuple[str, ...], str]:
    """
    既存notificationシートから通知状態を取得する。

    データ内容が同一なら、SUCCESSなどを維持する。
    内容が変わった場合は新しい行となるためPENDINGになる。
    """
    if worksheet.max_row < 2:
        return {}

    current_headers = [
        normalize_text(cell.value)
        for cell in worksheet[1][:len(columns)]
    ]

    if current_headers != columns:
        return {}

    status_map: dict[tuple[str, ...], str] = {}

    data_columns = [
        column
        for column in columns
        if column != "notification_status"
    ]

    for row_values in worksheet.iter_rows(
        min_row=2,
        max_col=len(columns),
        values_only=True,
    ):
        row = {
            column: row_values[index]
            for index, column in enumerate(columns)
        }

        status = normalize_text(
            row.get("notification_status")
        ).upper()

        if status not in VALID_NOTIFICATION_STATUSES:
            continue

        key = build_notification_key(
            row=row,
            data_columns=data_columns,
        )

        status_map[key] = status

    return status_map


def build_notification_key(
    row: dict[str, object],
    data_columns: list[str],
) -> tuple[str, ...]:
    """通知内容から既存状態照合用のキーを作成する。"""
    return tuple(
        normalize_text(row.get(column))
        for column in data_columns
    )


def apply_existing_status(
    row: dict[str, object],
    existing_statuses: dict[tuple[str, ...], str],
    data_columns: list[str],
) -> dict[str, object]:
    """同一内容の既存通知があれば、その状態を引き継ぐ。"""
    key = build_notification_key(
        row=row,
        data_columns=data_columns,
    )

    row["notification_status"] = existing_statuses.get(
        key,
        PENDING,
    )

    return row


# ============================================================
# 即時通知
# ============================================================

def calculate_realtime_notifications(
    raw_rows: list[dict[str, object]],
    existing_statuses: dict[tuple[str, ...], str],
) -> list[dict[str, object]]:
    """realtime_sheetから即時通知データを作成する。"""
    unique_rows = remove_exact_duplicate_rows(
        rows=raw_rows,
        columns=(
            "timestamp",
            "device_id",
            "animal_type",
            "confidence",
            "action_triggered",
            "stay_duration",
        ),
    )

    notification_rows: list[dict[str, object]] = []

    data_columns = [
        "timestamp",
        "device_id",
        "animal_type",
        "confidence",
        "action_triggered",
        "stay_duration",
    ]

    for raw_row in unique_rows:
        timestamp_value = raw_row.get("timestamp")

        if isinstance(timestamp_value, datetime):
            timestamp_value = timestamp_value.strftime(
                "%Y-%m-%d %H:%M:%S.%f"
            ).rstrip("0").rstrip(".")

        notification_row: dict[str, object] = {
            "timestamp": timestamp_value,
            "device_id": normalize_text(
                raw_row.get("device_id")
            ),
            "animal_type": normalize_text(
                raw_row.get("animal_type")
            ),
            "confidence": round(
                to_float(raw_row.get("confidence")),
                4,
            ),
            "action_triggered": normalize_text(
                raw_row.get("action_triggered")
            ),
            "stay_duration": round(
                to_float(raw_row.get("stay_duration")),
                3,
            ),
        }

        notification_rows.append(
            apply_existing_status(
                row=notification_row,
                existing_statuses=existing_statuses,
                data_columns=data_columns,
            )
        )

    return notification_rows


# ============================================================
# 日次通知
# ============================================================

def get_highest_alert_level(
    levels: list[str],
) -> str:
    """警戒レベル一覧から最も高い値を返す。"""
    normalized_levels = [
        normalize_text(level).upper()
        for level in levels
    ]

    valid_levels = [
        level
        for level in normalized_levels
        if level in ALERT_LEVEL_PRIORITY
    ]

    if not valid_levels:
        return "UNKNOWN"

    return max(
        valid_levels,
        key=lambda level: ALERT_LEVEL_PRIORITY[level],
    )


def calculate_daily_notifications(
    raw_rows: list[dict[str, object]],
    existing_statuses: dict[tuple[str, ...], str],
) -> list[dict[str, object]]:
    """daily_sheetを日付単位で集約する。"""
    latest_rows = keep_latest_rows(
        rows=raw_rows,
        key_columns=("date", "animal"),
    )

    grouped_rows: dict[
        str,
        list[dict[str, object]],
    ] = {}

    for raw_row in latest_rows:
        target_date = normalize_date_value(
            raw_row.get("date")
        )

        if not target_date:
            continue

        grouped_rows.setdefault(
            target_date,
            [],
        ).append(raw_row)

    notification_rows: list[dict[str, object]] = []

    data_columns = [
        "date",
        "total_detection_count",
        "boar_count",
        "monkey_count",
        "average_stay_time",
        "night_alert_level",
    ]

    for target_date in sorted(grouped_rows):
        date_rows = grouped_rows[target_date]

        total_detection_count = 0
        boar_count = 0
        monkey_count = 0
        weighted_stay_time = 0.0
        alert_levels: list[str] = []

        for raw_row in date_rows:
            animal = normalize_text(
                raw_row.get("animal")
            )

            total_count = to_int(
                raw_row.get("total_count")
            )

            average_stay_time = to_float(
                raw_row.get("average_stay_time")
            )

            total_detection_count += total_count

            weighted_stay_time += (
                average_stay_time
                * total_count
            )

            if animal in {"イノシシ", "boar"}:
                boar_count += total_count

            elif animal in {"サル", "monkey"}:
                monkey_count += total_count

            alert_levels.append(
                normalize_text(
                    raw_row.get("alert_level")
                )
            )

        if total_detection_count > 0:
            overall_average_stay_time = (
                weighted_stay_time
                / total_detection_count
            )
        else:
            overall_average_stay_time = 0.0

        notification_row: dict[str, object] = {
            "date": target_date,
            "total_detection_count": total_detection_count,
            "boar_count": boar_count,
            "monkey_count": monkey_count,
            "average_stay_time": round(
                overall_average_stay_time,
                3,
            ),
            "night_alert_level": get_highest_alert_level(
                alert_levels
            ),
        }

        notification_rows.append(
            apply_existing_status(
                row=notification_row,
                existing_statuses=existing_statuses,
                data_columns=data_columns,
            )
        )

    return notification_rows


# ============================================================
# 週次通知
# ============================================================

def calculate_weekly_notifications(
    raw_rows: list[dict[str, object]],
    existing_statuses: dict[tuple[str, ...], str],
) -> list[dict[str, object]]:
    """weekly_sheetから週次通知用データを作成する。"""
    latest_rows = keep_latest_rows(
        rows=raw_rows,
        key_columns=("week", "animal"),
    )

    notification_rows: list[dict[str, object]] = []

    data_columns = [
        "week",
        "animal",
        "peak_hour",
        "peak_day",
        "familiarity_score",
        "trap_score",
        "level",
    ]

    for raw_row in latest_rows:
        notification_row: dict[str, object] = {
            "week": normalize_text(
                raw_row.get("week")
            ),
            "animal": normalize_text(
                raw_row.get("animal")
            ),
            "peak_hour": to_int(
                raw_row.get("peak_hour")
            ),
            "peak_day": normalize_text(
                raw_row.get("peak_day")
            ),
            "familiarity_score": round(
                to_float(
                    raw_row.get("familiarity_score")
                ),
                4,
            ),
            "trap_score": round(
                to_float(
                    raw_row.get("trap_score")
                ),
                4,
            ),
            "level": normalize_text(
                raw_row.get("level")
            ).upper(),
        }

        notification_rows.append(
            apply_existing_status(
                row=notification_row,
                existing_statuses=existing_statuses,
                data_columns=data_columns,
            )
        )

    notification_rows.sort(
        key=lambda row: (
            normalize_text(row.get("week")),
            normalize_text(row.get("animal")),
        )
    )

    return notification_rows


# ============================================================
# 月次通知
# ============================================================

def calculate_peak_hour_for_month(
    realtime_rows: list[dict[str, object]],
    target_month: str,
    device_id: str,
) -> int | None:
    """対象月・対象カメラの出没ピーク時間を計算する。"""
    hour_counter: Counter[int] = Counter()

    for raw_row in realtime_rows:
        row_device_id = normalize_text(
            raw_row.get("device_id")
        )

        if row_device_id != device_id:
            continue

        timestamp = parse_timestamp(
            raw_row.get("timestamp")
        )

        if timestamp is None:
            continue

        if timestamp.strftime("%Y-%m") != target_month:
            continue

        hour_counter[timestamp.hour] += 1

    if not hour_counter:
        return None

    # 同件数の場合は早い時間を採用する。
    return min(
        hour_counter,
        key=lambda hour: (
            -hour_counter[hour],
            hour,
        ),
    )


def month_difference(
    previous_month: str,
    current_month: str,
) -> int | None:
    """2つのYYYY-MM間の月数差を返す。"""
    try:
        previous_date = datetime.strptime(
            previous_month,
            "%Y-%m",
        )

        current_date = datetime.strptime(
            current_month,
            "%Y-%m",
        )

    except ValueError:
        return None

    return (
        (current_date.year - previous_date.year) * 12
        + current_date.month
        - previous_date.month
    )


def calculate_previous_month_comparison(
    previous_total: int | None,
    previous_month: str | None,
    current_total: int,
    current_month: str,
) -> float | None:
    """前月比を計算する。"""
    if previous_total is None or previous_month is None:
        return None

    # 直前行ではなく、本当に1か月前か確認する。
    if month_difference(
        previous_month,
        current_month,
    ) != 1:
        return None

    if previous_total <= 0:
        return None

    comparison = (
        (current_total - previous_total)
        / previous_total
        * 100
    )

    return round(comparison, 2)


def calculate_action_effectiveness(
    realtime_rows: list[dict[str, object]],
    target_month: str,
    device_id: str,
) -> str | None:
    """
    撃退アクションの有効度を返す。

    現在のバックエンド仕様には正式な判定式がないため、
    根拠のない高・中・低判定は行わない。

    実行アクションが存在しても現段階ではNoneを返す。
    """
    has_executed_action = False

    no_action_values = {
        "",
        "なし",
        "none",
        "no_action",
        "not_triggered",
    }

    for raw_row in realtime_rows:
        if normalize_text(
            raw_row.get("device_id")
        ) != device_id:
            continue

        timestamp = parse_timestamp(
            raw_row.get("timestamp")
        )

        if timestamp is None:
            continue

        if timestamp.strftime("%Y-%m") != target_month:
            continue

        action = normalize_text(
            raw_row.get("action_triggered")
        ).lower()

        if action not in no_action_values:
            has_executed_action = True
            break

    if not has_executed_action:
        return None

    # 正式な有効度計算式が未定義。
    return None


def calculate_monthly_notifications(
    monthly_rows: list[dict[str, object]],
    realtime_rows: list[dict[str, object]],
    existing_statuses: dict[tuple[str, ...], str],
) -> list[dict[str, object]]:
    """monthly_sheetから月次通知データを作成する。"""
    latest_rows = keep_latest_rows(
        rows=monthly_rows,
        key_columns=("month", "device_id"),
    )

    normalized_rows: list[dict[str, object]] = []

    for raw_row in latest_rows:
        normalized_rows.append({
            "month": normalize_month_value(
                raw_row.get("month")
            ),
            "device_id": normalize_text(
                raw_row.get("device_id")
            ),
            "total_count": to_int(
                raw_row.get("total_count")
            ),
            "monkey_count": to_int(
                raw_row.get("monkey_count")
            ),
            "boar_count": to_int(
                raw_row.get("boar_count")
            ),
            "trap_score": to_float(
                raw_row.get("trap_score")
            ),
        })

    normalized_rows.sort(
        key=lambda row: (
            normalize_text(row.get("device_id")),
            normalize_text(row.get("month")),
        )
    )

    previous_values: dict[
        str,
        tuple[str, int],
    ] = {}

    notification_rows: list[dict[str, object]] = []

    data_columns = [
        "month",
        "device_id",
        "monthly_total_count",
        "boar_ratio",
        "monkey_ratio",
        "comparison_previous_month",
        "monthly_peak_hour",
        "action_effectiveness",
        "trap_score",
    ]

    for row in normalized_rows:
        target_month = normalize_text(
            row.get("month")
        )

        device_id = normalize_text(
            row.get("device_id")
        )

        total_count = to_int(
            row.get("total_count")
        )

        monkey_count = to_int(
            row.get("monkey_count")
        )

        boar_count = to_int(
            row.get("boar_count")
        )

        if total_count > 0:
            boar_ratio = round(
                boar_count / total_count * 100,
                2,
            )

            monkey_ratio = round(
                monkey_count / total_count * 100,
                2,
            )

        else:
            boar_ratio = 0.0
            monkey_ratio = 0.0

        previous_month: str | None = None
        previous_total: int | None = None

        if device_id in previous_values:
            previous_month, previous_total = (
                previous_values[device_id]
            )

        comparison_previous_month = (
            calculate_previous_month_comparison(
                previous_total=previous_total,
                previous_month=previous_month,
                current_total=total_count,
                current_month=target_month,
            )
        )

        monthly_peak_hour = calculate_peak_hour_for_month(
            realtime_rows=realtime_rows,
            target_month=target_month,
            device_id=device_id,
        )

        action_effectiveness = calculate_action_effectiveness(
            realtime_rows=realtime_rows,
            target_month=target_month,
            device_id=device_id,
        )

        notification_row: dict[str, object] = {
            "month": target_month,
            "device_id": device_id,
            "monthly_total_count": total_count,
            "boar_ratio": boar_ratio,
            "monkey_ratio": monkey_ratio,
            "comparison_previous_month": (
                comparison_previous_month
            ),
            "monthly_peak_hour": monthly_peak_hour,
            "action_effectiveness": action_effectiveness,
            "trap_score": round(
                to_float(row.get("trap_score")),
                4,
            ),
        }

        notification_rows.append(
            apply_existing_status(
                row=notification_row,
                existing_statuses=existing_statuses,
                data_columns=data_columns,
            )
        )

        previous_values[device_id] = (
            target_month,
            total_count,
        )

    return notification_rows


# ============================================================
# 年次通知
# ============================================================

def calculate_yearly_notifications(
    raw_rows: list[dict[str, object]],
    existing_statuses: dict[tuple[str, ...], str],
) -> list[dict[str, object]]:
    """yearly_sheetから年次通知データを作成する。"""
    latest_rows = keep_latest_rows(
        rows=raw_rows,
        key_columns=("year", "device_id"),
    )

    notification_rows: list[dict[str, object]] = []

    data_columns = [
        "year",
        "device_id",
        "total_count",
        "monkey_count",
        "boar_count",
        "trap_score",
    ]

    for raw_row in latest_rows:
        notification_row: dict[str, object] = {
            "year": normalize_year_value(
                raw_row.get("year")
            ),
            "device_id": normalize_text(
                raw_row.get("device_id")
            ),
            "total_count": to_int(
                raw_row.get("total_count")
            ),
            "monkey_count": to_int(
                raw_row.get("monkey_count")
            ),
            "boar_count": to_int(
                raw_row.get("boar_count")
            ),
            "trap_score": round(
                to_float(
                    raw_row.get("trap_score")
                ),
                4,
            ),
        }

        notification_rows.append(
            apply_existing_status(
                row=notification_row,
                existing_statuses=existing_statuses,
                data_columns=data_columns,
            )
        )

    notification_rows.sort(
        key=lambda row: (
            to_int(row.get("year")),
            normalize_text(row.get("device_id")),
        )
    )

    return notification_rows


# ============================================================
# 全通知計算
# ============================================================

def calculate_all_notifications(
    workbook_path: Path,
) -> dict[str, list[dict[str, object]]]:
    """
    notification_database.xlsxを読み込み、
    5種類のnotificationシート用データを返す。

    この関数ではExcelへの書き込みは行わない。
    """
    workbook_path = Path(workbook_path).resolve()

    if not workbook_path.is_file():
        raise FileNotFoundError(
            "notification_database.xlsxが見つかりません: "
            f"{workbook_path}"
        )

    LOGGER.info(
        "通知用データの計算元Excelを読み込みます: %s",
        workbook_path,
    )

    workbook = load_workbook(
        workbook_path,
        data_only=True,
        read_only=False,
    )

    try:
        validate_required_sheets(workbook)

        raw_data: dict[str, list[dict[str, object]]] = {}

        for sheet_name, required_columns in (
            RAW_REQUIRED_COLUMNS.items()
        ):
            raw_data[sheet_name] = read_worksheet_rows(
                worksheet=workbook[sheet_name],
                required_columns=required_columns,
            )

            LOGGER.info(
                "%sから%d件読み込みました。",
                sheet_name,
                len(raw_data[sheet_name]),
            )

        existing_statuses: dict[
            str,
            dict[tuple[str, ...], str],
        ] = {}

        for (
            notification_sheet_name,
            notification_columns,
        ) in NOTIFICATION_COLUMNS.items():
            existing_statuses[notification_sheet_name] = (
                load_existing_statuses(
                    worksheet=workbook[
                        notification_sheet_name
                    ],
                    columns=notification_columns,
                )
            )

        results = {
            "realtime_notification": (
                calculate_realtime_notifications(
                    raw_rows=raw_data["realtime_sheet"],
                    existing_statuses=existing_statuses[
                        "realtime_notification"
                    ],
                )
            ),
            "daily_notification": (
                calculate_daily_notifications(
                    raw_rows=raw_data["daily_sheet"],
                    existing_statuses=existing_statuses[
                        "daily_notification"
                    ],
                )
            ),
            "weekly_notification": (
                calculate_weekly_notifications(
                    raw_rows=raw_data["weekly_sheet"],
                    existing_statuses=existing_statuses[
                        "weekly_notification"
                    ],
                )
            ),
            "monthly_notification": (
                calculate_monthly_notifications(
                    monthly_rows=raw_data["monthly_sheet"],
                    realtime_rows=raw_data["realtime_sheet"],
                    existing_statuses=existing_statuses[
                        "monthly_notification"
                    ],
                )
            ),
            "yearly_notification": (
                calculate_yearly_notifications(
                    raw_rows=raw_data["yearly_sheet"],
                    existing_statuses=existing_statuses[
                        "yearly_notification"
                    ],
                )
            ),
        }

    finally:
        workbook.close()

    for sheet_name, rows in results.items():
        LOGGER.info(
            "%s用に%d件作成しました。",
            sheet_name,
            len(rows),
        )

    return results


# ============================================================
# 単独確認
# ============================================================

def main() -> None:
    """Calculator単体で計算結果だけ確認する。"""
    import argparse
    import json

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description=(
            "notification_database.xlsxから"
            "通知用データを計算します。"
        )
    )

    parser.add_argument(
        "workbook_path",
        type=Path,
        help="notification_database.xlsxのパス",
    )

    args = parser.parse_args()

    results = calculate_all_notifications(
        workbook_path=args.workbook_path,
    )

    print(
        json.dumps(
            results,
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    try:
        main()

    except (
        FileNotFoundError,
        NotificationCalculationError,
        ValueError,
    ) as error:
        LOGGER.error("%s", error)
        raise SystemExit(1)

    except Exception:
        LOGGER.exception(
            "通知用データの計算に失敗しました。"
        )
        raise SystemExit(1)