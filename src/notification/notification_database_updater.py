from __future__ import annotations

import csv
import logging
import os
import tempfile
from pathlib import Path
from typing import Any

from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger(__name__)


# ============================================================
# パス設定
# ============================================================

# project-root/src/notification/
NOTIFICATION_DIR = Path(__file__).resolve().parent

# project-root/src/
SRC_DIR = NOTIFICATION_DIR.parent

# project-root/src/backend/
BACKEND_DIR = SRC_DIR / "backend"

# project-root/src/notification/result/
RESULT_DIR = NOTIFICATION_DIR / "result"

# 既存のExcelファイル
DATABASE_XLSX = RESULT_DIR / "notification_database.xlsx"

# Excel同時アクセス防止用
DATABASE_LOCK_FILE = RESULT_DIR / ".notification_database.xlsx.lock"

LOCK_TIMEOUT_SECONDS = 30


# ============================================================
# CSVとrawシートの対応
# ============================================================

SOURCE_SETTINGS: dict[str, dict[str, Any]] = {
    "detections.csv": {
        "path": BACKEND_DIR / "detections.csv",
        "sheet_name": "realtime_sheet",
        "columns": [
            "timestamp",
            "device_id",
            "animal_type",
            "confidence",
            "action_triggered",
            "stay_duration",
        ],
    },
    "daily_analysis.csv": {
        "path": BACKEND_DIR / "daily_analysis.csv",
        "sheet_name": "daily_sheet",
        "columns": [
            "date",
            "animal",
            "total_count",
            "average_stay_time",
            "alert_level",
        ],
    },
    "weekly_analysis.csv": {
        "path": BACKEND_DIR / "weekly_analysis.csv",
        "sheet_name": "weekly_sheet",
        "columns": [
            "week",
            "animal",
            "peak_hour",
            "peak_day",
            "familiarity_score",
            "trap_score",
            "level",
        ],
    },
    "monthly_analysis.csv": {
        "path": BACKEND_DIR / "monthly_analysis.csv",
        "sheet_name": "monthly_sheet",
        "columns": [
            "month",
            "device_id",
            "total_count",
            "monkey_count",
            "boar_count",
            "trap_score",
            "rank",
        ],
    },
    "yearly_analysis.csv": {
        "path": BACKEND_DIR / "yearly_analysis.csv",
        "sheet_name": "yearly_sheet",
        "columns": [
            "year",
            "device_id",
            "total_count",
            "monkey_count",
            "boar_count",
            "trap_score",
            "rank",
        ],
    },
}


# Updaterでは存在確認だけ行い、中身は変更しない
NOTIFICATION_SHEETS = [
    "realtime_notification",
    "daily_notification",
    "weekly_notification",
    "monthly_notification",
    "yearly_notification",
]


# ============================================================
# 入力検証
# ============================================================

def validate_required_files() -> None:
    """
    バックエンドCSVとnotification_database.xlsxの存在を確認する。

    Excelは自動生成しない。
    """
    if not BACKEND_DIR.is_dir():
        raise FileNotFoundError(
            f"backendディレクトリがありません: {BACKEND_DIR}"
        )

    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if not DATABASE_XLSX.is_file():
        raise FileNotFoundError(
            "notification_database.xlsxがありません。\n"
            f"配置先: {DATABASE_XLSX}"
        )

    missing_csv_files: list[Path] = []

    for setting in SOURCE_SETTINGS.values():
        csv_path = Path(setting["path"])

        if not csv_path.is_file():
            missing_csv_files.append(csv_path)

    if missing_csv_files:
        missing_text = "\n".join(
            f"- {path}"
            for path in missing_csv_files
        )

        raise FileNotFoundError(
            "次の入力CSVがありません。\n"
            f"{missing_text}"
        )


def validate_workbook_sheets(workbook: Any) -> None:
    """
    rawシート5枚とnotificationシート5枚の存在を確認する。

    不足シートは自動生成しない。
    """
    raw_sheets = [
        str(setting["sheet_name"])
        for setting in SOURCE_SETTINGS.values()
    ]

    required_sheets = raw_sheets + NOTIFICATION_SHEETS

    missing_sheets = [
        sheet_name
        for sheet_name in required_sheets
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        missing_text = "\n".join(
            f"- {sheet_name}"
            for sheet_name in missing_sheets
        )

        raise ValueError(
            "notification_database.xlsxに"
            "必要なシートがありません。\n"
            f"{missing_text}"
        )


# ============================================================
# CSV読込
# ============================================================

def read_csv_file(
    csv_path: Path,
    expected_columns: list[str],
) -> list[dict[str, str]]:
    """
    CSVを読み込み、ヘッダー名と項目順を厳密に検証する。

    UTF-8およびUTF-8 BOM付きCSVに対応する。
    """
    try:
        with csv_path.open(
            mode="r",
            encoding="utf-8-sig",
            newline="",
        ) as csv_file:
            reader = csv.DictReader(csv_file)

            if reader.fieldnames is None:
                raise ValueError(
                    f"CSVにヘッダーがありません: {csv_path}"
                )

            actual_columns = [
                column.strip()
                for column in reader.fieldnames
            ]

            if actual_columns != expected_columns:
                missing_columns = [
                    column
                    for column in expected_columns
                    if column not in actual_columns
                ]

                extra_columns = [
                    column
                    for column in actual_columns
                    if column not in expected_columns
                ]

                raise ValueError(
                    "CSV項目が仕様と一致していません。\n"
                    f"対象ファイル: {csv_path}\n"
                    f"期待する項目: {expected_columns}\n"
                    f"実際の項目: {actual_columns}\n"
                    f"不足項目: {missing_columns}\n"
                    f"余分な項目: {extra_columns}"
                )

            rows: list[dict[str, str]] = []

            for line_number, raw_row in enumerate(
                reader,
                start=2,
            ):
                normalized_row: dict[str, str] = {}

                for key, value in raw_row.items():
                    if key is None:
                        continue

                    normalized_key = key.strip()

                    normalized_row[normalized_key] = (
                        value.strip()
                        if isinstance(value, str)
                        else ""
                    )

                # 完全な空行は無視
                if all(
                    normalized_row.get(column, "") == ""
                    for column in expected_columns
                ):
                    continue

                missing_values = [
                    column
                    for column in expected_columns
                    if column not in normalized_row
                ]

                if missing_values:
                    raise ValueError(
                        f"{csv_path.name}の{line_number}行目に"
                        f"不足項目があります: {missing_values}"
                    )

                rows.append({
                    column: normalized_row.get(column, "")
                    for column in expected_columns
                })

    except UnicodeError as error:
        raise ValueError(
            f"CSVをUTF-8として読み込めません: {csv_path}"
        ) from error

    LOGGER.info(
        "%sから%d件読み込みました。",
        csv_path.name,
        len(rows),
    )

    return rows


# ============================================================
# Excel向け型変換
# ============================================================

def convert_excel_value(
    column_name: str,
    value: str,
) -> Any:
    """CSV文字列をExcel向けの型へ変換する。"""
    if value == "":
        return None

    integer_columns = {
        "total_count",
        "monkey_count",
        "boar_count",
        "year",
        "rank",
        "peak_hour",
    }

    float_columns = {
        "confidence",
        "stay_duration",
        "average_stay_time",
        "familiarity_score",
        "trap_score",
    }

    if column_name in integer_columns:
        try:
            return int(float(value))
        except ValueError as error:
            raise ValueError(
                f"{column_name}には整数が必要です。"
                f"入力値={value}"
            ) from error

    if column_name in float_columns:
        try:
            return float(value)
        except ValueError as error:
            raise ValueError(
                f"{column_name}には数値が必要です。"
                f"入力値={value}"
            ) from error

    return value


# ============================================================
# rawシート更新
# ============================================================

def replace_raw_sheet_data(
    worksheet: Worksheet,
    columns: list[str],
    rows: list[dict[str, str]],
) -> None:
    """
    rawシートを最新CSVの内容へ置き換える。

    notificationシートには触れない。
    """
    if worksheet.max_row > 0:
        worksheet.delete_rows(
            idx=1,
            amount=worksheet.max_row,
        )

    worksheet.append(columns)

    for row in rows:
        worksheet.append([
            convert_excel_value(
                column_name=column,
                value=row.get(column, ""),
            )
            for column in columns
        ])

    format_raw_sheet(
        worksheet=worksheet,
        columns=columns,
    )


def format_raw_sheet(
    worksheet: Worksheet,
    columns: list[str],
) -> None:
    """rawシートの表示形式を設定する。"""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    column_positions = {
        column: index + 1
        for index, column in enumerate(columns)
    }

    set_number_format(
        worksheet=worksheet,
        column_positions=column_positions,
        column_name="confidence",
        number_format="0.0000",
    )

    for column_name in [
        "stay_duration",
        "average_stay_time",
        "familiarity_score",
        "trap_score",
    ]:
        set_number_format(
            worksheet=worksheet,
            column_positions=column_positions,
            column_name=column_name,
            number_format="0.00",
        )

    adjust_column_widths(worksheet)


def set_number_format(
    worksheet: Worksheet,
    column_positions: dict[str, int],
    column_name: str,
    number_format: str,
) -> None:
    """指定列へExcelの数値表示形式を設定する。"""
    column_index = column_positions.get(column_name)

    if column_index is None:
        return

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row_number,
            column=column_index,
        ).number_format = number_format


def adjust_column_widths(
    worksheet: Worksheet,
) -> None:
    """セル内容に合わせて列幅を調整する。"""
    for column_cells in worksheet.columns:
        maximum_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(cell.value)),
            )

        column_letter = column_cells[0].column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 2, 12),
            35,
        )


# ============================================================
# Excel保存
# ============================================================

def save_workbook_atomically(
    workbook: Any,
) -> None:
    """
    一時ファイルへ保存してから本体Excelと置き換える。

    保存失敗時に既存Excelを破損させない。
    """
    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix="notification_database_",
        suffix=".xlsx",
        dir=RESULT_DIR,
    )

    os.close(file_descriptor)

    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)

        os.replace(
            temporary_path,
            DATABASE_XLSX,
        )

    finally:
        if temporary_path.exists():
            temporary_path.unlink()


# ============================================================
# 全体処理
# ============================================================

def update_notification_database() -> Path:
    """
    バックエンドの5つのCSVを読み込み、
    notification_database.xlsxのrawシートだけを更新する。

    Launcherから呼び出される公開関数。

    Returns
    -------
    Path
        更新したnotification_database.xlsxの絶対パス。
    """
    validate_required_files()

    # Excelを開く前に全CSVを検証する。
    # 1ファイルでも不正ならExcelを更新しない。
    source_data: dict[str, list[dict[str, str]]] = {}

    for filename, setting in SOURCE_SETTINGS.items():
        csv_path = Path(setting["path"])

        source_data[filename] = read_csv_file(
            csv_path=csv_path,
            expected_columns=list(setting["columns"]),
        )

    lock = FileLock(
        str(DATABASE_LOCK_FILE),
        timeout=LOCK_TIMEOUT_SECONDS,
    )

    try:
        with lock:
            workbook = load_workbook(DATABASE_XLSX)

            try:
                validate_workbook_sheets(workbook)

                for filename, setting in SOURCE_SETTINGS.items():
                    sheet_name = str(setting["sheet_name"])
                    columns = list(setting["columns"])

                    worksheet = workbook[sheet_name]

                    replace_raw_sheet_data(
                        worksheet=worksheet,
                        columns=columns,
                        rows=source_data[filename],
                    )

                    LOGGER.info(
                        "%sを%sへ反映しました。",
                        filename,
                        sheet_name,
                    )

                save_workbook_atomically(workbook)

            finally:
                workbook.close()

    except Timeout as error:
        raise TimeoutError(
            "notification_database.xlsxのロックを"
            f"{LOCK_TIMEOUT_SECONDS}秒以内に取得できませんでした。"
        ) from error

    LOGGER.info(
        "notification_database.xlsxのrawシートを更新しました: %s",
        DATABASE_XLSX,
    )

    LOGGER.info(
        "notificationシート5枚は変更していません。"
    )

    return DATABASE_XLSX.resolve()


# ============================================================
# 単独実行
# ============================================================

def main() -> None:
    """Updaterだけを単独で実行する。"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    workbook_path = update_notification_database()

    LOGGER.info(
        "更新完了: %s",
        workbook_path,
    )


if __name__ == "__main__":
    try:
        main()

    except PermissionError as error:
        LOGGER.error(
            "notification_database.xlsxへのアクセスが拒否されました。"
            "Excelで開いている場合は閉じてください。"
        )
        raise SystemExit(1) from error

    except (
        FileNotFoundError,
        TimeoutError,
        ValueError,
    ) as error:
        LOGGER.error("%s", error)
        raise SystemExit(1) from error

    except Exception as error:
        LOGGER.exception(
            "rawシート更新処理に失敗しました。"
        )
        raise SystemExit(1) from error