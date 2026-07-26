from __future__ import annotations

import argparse
import json
import logging
import os
import tempfile
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


LOGGER = logging.getLogger(__name__)

# ============================================================
# パス設定
# ============================================================

NOTIFICATION_DIR = Path(__file__).resolve().parent
SRC_DIR = NOTIFICATION_DIR.parent
PROJECT_ROOT = SRC_DIR.parent

ROOT_ENV_PATH = PROJECT_ROOT / ".env"
SRC_ENV_PATH = SRC_DIR / ".env"

LOCK_TIMEOUT_SECONDS = 30
REQUEST_TIMEOUT_SECONDS = 15
MAX_RETRY_COUNT = 3
RETRY_WAIT_SECONDS = 2.0

SLACK_WEBHOOK_ENV_NAME = "SLACK_WEBHOOK_URL"

PENDING = "PENDING"
SUCCESS = "SUCCESS"
FAILED = "FAILED"
SKIPPED = "SKIPPED"

SUPPORTED_STATUSES = {
    PENDING,
    SUCCESS,
    FAILED,
    SKIPPED,
}

STATUS_FILL_COLORS = {
    PENDING: "FFF2CC",
    SUCCESS: "C6E0B4",
    FAILED: "F4CCCC",
    SKIPPED: "D9E1F2",
}

NOTIFICATION_SHEETS = (
    "realtime_notification",
    "daily_notification",
    "weekly_notification",
    "monthly_notification",
    "yearly_notification",
)


# ============================================================
# 結果型・例外
# ============================================================

class NotificationSenderError(RuntimeError):
    """Slack通知処理に失敗した場合の例外。"""


@dataclass(frozen=True)
class SlackSendResult:
    success: bool
    status_code: int | None
    response_text: str
    error_message: str | None = None


# ============================================================
# 環境変数
# ============================================================

def load_environment() -> Path | None:
    """
    .envを読み込む。

    優先順位:
    1. 既存のOS環境変数
    2. プロジェクトルート/.env
    3. src/.env

    load_dotenv()はoverride=Falseで使用するため、
    既存のOS環境変数を上書きしない。
    """
    if ROOT_ENV_PATH.is_file():
        load_dotenv(
            dotenv_path=ROOT_ENV_PATH,
            override=False,
        )
        LOGGER.info(".envを読み込みました: %s", ROOT_ENV_PATH)
        return ROOT_ENV_PATH

    if SRC_ENV_PATH.is_file():
        load_dotenv(
            dotenv_path=SRC_ENV_PATH,
            override=False,
        )
        LOGGER.warning(
            "src/.envを読み込みました。"
            "推奨配置はプロジェクトルート/.envです: %s",
            SRC_ENV_PATH,
        )
        return SRC_ENV_PATH

    LOGGER.warning(
        ".envが見つかりません。OS環境変数から%sを取得します。",
        SLACK_WEBHOOK_ENV_NAME,
    )
    return None


def get_slack_webhook_url() -> str:
    """Slack Incoming Webhook URLを環境変数から取得する。"""
    load_environment()

    webhook_url = os.getenv(
        SLACK_WEBHOOK_ENV_NAME,
        "",
    ).strip()

    if not webhook_url:
        raise NotificationSenderError(
            f"環境変数{SLACK_WEBHOOK_ENV_NAME}が設定されていません。"
        )

    if not (
        webhook_url.startswith("https://hooks.slack.com/")
        or webhook_url.startswith("https://hooks.slack-gov.com/")
    ):
        raise NotificationSenderError(
            "SLACK_WEBHOOK_URLの形式がSlack Incoming Webhookではありません。"
        )

    return webhook_url


# ============================================================
# Excel読込
# ============================================================

def normalize_text(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def read_header_map(
    worksheet: Worksheet,
) -> dict[str, int]:
    """1行目のヘッダー名と列番号を取得する。"""
    header_map: dict[str, int] = {}

    for cell in worksheet[1]:
        header = normalize_text(cell.value)

        if header:
            header_map[header] = cell.column

    return header_map


def validate_workbook(workbook: Any) -> None:
    """必要な通知シートとnotification_status列を検証する。"""
    missing_sheets = [
        sheet_name
        for sheet_name in NOTIFICATION_SHEETS
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise NotificationSenderError(
            "notification_database.xlsxに必要な通知シートがありません。\n"
            + "\n".join(
                f"- {sheet_name}"
                for sheet_name in missing_sheets
            )
        )

    for sheet_name in NOTIFICATION_SHEETS:
        header_map = read_header_map(
            workbook[sheet_name]
        )

        if "notification_status" not in header_map:
            raise NotificationSenderError(
                f"{sheet_name}にnotification_status列がありません。"
            )


def worksheet_row_to_dict(
    worksheet: Worksheet,
    row_number: int,
    header_map: dict[str, int],
) -> dict[str, object]:
    """指定行を辞書へ変換する。"""
    return {
        header: worksheet.cell(
            row=row_number,
            column=column_number,
        ).value
        for header, column_number in header_map.items()
    }


# ============================================================
# Slackメッセージ生成
# ============================================================

def format_optional_value(
    value: object,
    suffix: str = "",
) -> str:
    text = normalize_text(value)

    if not text or text == "N/A":
        return "N/A"

    return f"{text}{suffix}"


def build_realtime_message(
    row: dict[str, object],
) -> str:
    return "\n".join([
        "【害獣出没リアルタイム速報】",
        f"・検知日時: {format_optional_value(row.get('timestamp'))}",
        f"・カメラID: {format_optional_value(row.get('device_id'))}",
        f"・動物の種類: {format_optional_value(row.get('animal_type'))}",
        f"・検知確信度: {format_optional_value(row.get('confidence'))}",
        f"・実行アクション: {format_optional_value(row.get('action_triggered'))}",
        f"・滞在時間: {format_optional_value(row.get('stay_duration'), '秒')}",
    ])


def build_daily_message(
    row: dict[str, object],
) -> str:
    return "\n".join([
        "【害獣出没 日次レポート】",
        f"・対象日: {format_optional_value(row.get('date'))}",
        f"・本日の総検知回数: {format_optional_value(row.get('total_detection_count'), '件')}",
        f"・イノシシ: {format_optional_value(row.get('boar_count'), '件')}",
        f"・サル: {format_optional_value(row.get('monkey_count'), '件')}",
        f"・平均滞在時間: {format_optional_value(row.get('average_stay_time'), '秒')}",
        f"・夜間警戒レベル: {format_optional_value(row.get('night_alert_level'))}",
    ])


def build_weekly_message(
    row: dict[str, object],
) -> str:
    return "\n".join([
        "【害獣出没 週次レポート】",
        f"・対象週: {format_optional_value(row.get('week'))}",
        f"・対象動物: {format_optional_value(row.get('animal'))}",
        f"・出没ピーク時間: {format_optional_value(row.get('peak_hour'), '時')}",
        f"・出没ピーク曜日: {format_optional_value(row.get('peak_day'))}",
        f"・慣れ度スコア: {format_optional_value(row.get('familiarity_score'))}",
        f"・罠スコア: {format_optional_value(row.get('trap_score'))}",
        f"・判定レベル: {format_optional_value(row.get('level'))}",
    ])


def build_monthly_message(
    row: dict[str, object],
) -> str:
    return "\n".join([
        "【害獣出没 月次レポート】",
        f"・対象月: {format_optional_value(row.get('month'))}",
        f"・カメラID: {format_optional_value(row.get('device_id'))}",
        f"・月間総検知数: {format_optional_value(row.get('monthly_total_count'), '件')}",
        f"・イノシシ割合: {format_optional_value(row.get('boar_ratio'), '%')}",
        f"・サル割合: {format_optional_value(row.get('monkey_ratio'), '%')}",
        f"・前月比: {format_optional_value(row.get('comparison_previous_month'), '%')}",
        f"・月間ピーク時間: {format_optional_value(row.get('monthly_peak_hour'), '時')}",
        f"・主要アクション有効度: {format_optional_value(row.get('action_effectiveness'))}",
        f"・罠スコア: {format_optional_value(row.get('trap_score'))}",
    ])


def build_yearly_message(
    row: dict[str, object],
) -> str:
    return "\n".join([
        "【害獣出没 年次レポート】",
        f"・対象年: {format_optional_value(row.get('year'))}",
        f"・カメラID: {format_optional_value(row.get('device_id'))}",
        f"・年間総検知数: {format_optional_value(row.get('total_count'), '件')}",
        f"・サル: {format_optional_value(row.get('monkey_count'), '件')}",
        f"・イノシシ: {format_optional_value(row.get('boar_count'), '件')}",
        f"・罠スコア: {format_optional_value(row.get('trap_score'))}",
    ])


MESSAGE_BUILDERS = {
    "realtime_notification": build_realtime_message,
    "daily_notification": build_daily_message,
    "weekly_notification": build_weekly_message,
    "monthly_notification": build_monthly_message,
    "yearly_notification": build_yearly_message,
}


# ============================================================
# Slack送信
# ============================================================

def post_to_slack(
    webhook_url: str,
    message: str,
) -> SlackSendResult:
    """Incoming WebhookへJSONメッセージをPOSTする。"""
    payload = json.dumps(
        {"text": message},
        ensure_ascii=False,
    ).encode("utf-8")

    request = urllib.request.Request(
        url=webhook_url,
        data=payload,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "User-Agent": "wildlife-notification-system/1.0",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(
            request,
            timeout=REQUEST_TIMEOUT_SECONDS,
        ) as response:
            status_code = response.getcode()
            response_text = response.read().decode(
                "utf-8",
                errors="replace",
            ).strip()

        success = (
            status_code == 200
            and response_text.lower() == "ok"
        )

        return SlackSendResult(
            success=success,
            status_code=status_code,
            response_text=response_text,
            error_message=(
                None
                if success
                else (
                    "Slackから正常応答を受け取れませんでした。"
                )
            ),
        )

    except urllib.error.HTTPError as error:
        response_text = error.read().decode(
            "utf-8",
            errors="replace",
        ).strip()

        return SlackSendResult(
            success=False,
            status_code=error.code,
            response_text=response_text,
            error_message=(
                f"Slack HTTPエラー: {error.code} "
                f"{response_text}"
            ),
        )

    except urllib.error.URLError as error:
        return SlackSendResult(
            success=False,
            status_code=None,
            response_text="",
            error_message=f"Slack接続エラー: {error.reason}",
        )

    except TimeoutError:
        return SlackSendResult(
            success=False,
            status_code=None,
            response_text="",
            error_message="Slack通知がタイムアウトしました。",
        )


def send_with_retry(
    webhook_url: str,
    message: str,
) -> SlackSendResult:
    """一時的な通信失敗に対して最大3回送信する。"""
    last_result: SlackSendResult | None = None

    for attempt in range(
        1,
        MAX_RETRY_COUNT + 1,
    ):
        result = post_to_slack(
            webhook_url=webhook_url,
            message=message,
        )

        if result.success:
            return result

        last_result = result

        # 400・403・404は設定・入力問題の可能性が高いため再試行しない。
        if result.status_code in {
            400,
            403,
            404,
        }:
            return result

        if attempt < MAX_RETRY_COUNT:
            LOGGER.warning(
                "Slack送信に失敗しました。"
                "%s秒後に再試行します。試行=%d/%d、理由=%s",
                RETRY_WAIT_SECONDS,
                attempt,
                MAX_RETRY_COUNT,
                result.error_message,
            )
            time.sleep(RETRY_WAIT_SECONDS)

    if last_result is None:
        raise NotificationSenderError(
            "Slack送信結果を取得できませんでした。"
        )

    return last_result


# ============================================================
# ステータス更新・保存
# ============================================================

def update_status_cell(
    worksheet: Worksheet,
    row_number: int,
    status_column: int,
    status: str,
) -> None:
    """notification_statusセルの値と色を更新する。"""
    if status not in SUPPORTED_STATUSES:
        raise NotificationSenderError(
            f"未対応の通知状態です: {status}"
        )

    cell = worksheet.cell(
        row=row_number,
        column=status_column,
    )

    cell.value = status

    fill_color = STATUS_FILL_COLORS.get(status)

    if fill_color:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color,
        )


def save_workbook_atomically(
    workbook: Any,
    workbook_path: Path,
) -> None:
    """一時ファイルへ保存後、元のExcelと置き換える。"""
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}_",
        suffix=".xlsx",
        dir=workbook_path.parent,
    )

    os.close(file_descriptor)

    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)
        os.replace(
            temporary_path,
            workbook_path,
        )

    finally:
        if temporary_path.exists():
            temporary_path.unlink()


# ============================================================
# 公開関数
# ============================================================

def send_pending_notifications(
    workbook_path: Path,
) -> dict[str, int]:
    """
    5つのnotificationシートからPENDING行だけをSlackへ送信する。

    送信成功:
        notification_status = SUCCESS

    送信失敗:
        notification_status = FAILED

    Returns
    -------
    dict[str, int]
        success・failed・skipped・totalの件数。
    """
    workbook_path = Path(workbook_path).resolve()

    if not workbook_path.is_file():
        raise FileNotFoundError(
            "notification_database.xlsxが見つかりません: "
            f"{workbook_path}"
        )

    webhook_url = get_slack_webhook_url()

    lock_path = workbook_path.parent / (
        f".{workbook_path.name}.lock"
    )

    lock = FileLock(
        str(lock_path),
        timeout=LOCK_TIMEOUT_SECONDS,
    )

    counts = {
        "success": 0,
        "failed": 0,
        "skipped": 0,
        "total": 0,
    }

    try:
        with lock:
            workbook = load_workbook(
                workbook_path,
                read_only=False,
                data_only=False,
            )

            try:
                validate_workbook(workbook)

                for sheet_name in NOTIFICATION_SHEETS:
                    worksheet = workbook[sheet_name]
                    header_map = read_header_map(
                        worksheet
                    )

                    status_column = header_map[
                        "notification_status"
                    ]

                    message_builder = MESSAGE_BUILDERS[
                        sheet_name
                    ]

                    for row_number in range(
                        2,
                        worksheet.max_row + 1,
                    ):
                        status = normalize_text(
                            worksheet.cell(
                                row=row_number,
                                column=status_column,
                            ).value
                        ).upper()

                        if not status:
                            counts["skipped"] += 1
                            continue

                        if status != PENDING:
                            counts["skipped"] += 1
                            continue

                        row = worksheet_row_to_dict(
                            worksheet=worksheet,
                            row_number=row_number,
                            header_map=header_map,
                        )

                        message = message_builder(row)

                        result = send_with_retry(
                            webhook_url=webhook_url,
                            message=message,
                        )

                        counts["total"] += 1

                        if result.success:
                            update_status_cell(
                                worksheet=worksheet,
                                row_number=row_number,
                                status_column=status_column,
                                status=SUCCESS,
                            )
                            counts["success"] += 1

                            LOGGER.info(
                                "Slack通知成功: %s 行%d",
                                sheet_name,
                                row_number,
                            )

                        else:
                            update_status_cell(
                                worksheet=worksheet,
                                row_number=row_number,
                                status_column=status_column,
                                status=FAILED,
                            )
                            counts["failed"] += 1

                            LOGGER.error(
                                "Slack通知失敗: %s 行%d、理由=%s",
                                sheet_name,
                                row_number,
                                result.error_message,
                            )

                save_workbook_atomically(
                    workbook=workbook,
                    workbook_path=workbook_path,
                )

            finally:
                workbook.close()

    except Timeout as error:
        raise TimeoutError(
            "notification_database.xlsxのロックを"
            f"{LOCK_TIMEOUT_SECONDS}秒以内に取得できませんでした。"
        ) from error

    LOGGER.info(
        "Slack通知処理が完了しました: %s",
        counts,
    )

    return counts


# ============================================================
# 単独実行
# ============================================================

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description=(
            "notification_database.xlsxのPENDING行を"
            "Slack Incoming Webhookへ送信します。"
        )
    )

    parser.add_argument(
        "workbook_path",
        type=Path,
        help="notification_database.xlsxのパス",
    )

    args = parser.parse_args()

    result = send_pending_notifications(
        workbook_path=args.workbook_path,
    )

    print(
        json.dumps(
            result,
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    try:
        main()

    except PermissionError:
        LOGGER.error(
            "notification_database.xlsxへ書き込めません。"
            "Excelで開いている場合は閉じてください。"
        )
        raise SystemExit(1)

    except (
        FileNotFoundError,
        NotificationSenderError,
        TimeoutError,
        ValueError,
    ) as error:
        LOGGER.error("%s", error)
        raise SystemExit(1)

    except KeyboardInterrupt:
        LOGGER.info("Slack通知処理を終了しました。")
        raise SystemExit(0)

    except Exception:
        LOGGER.exception(
            "Slack通知処理に失敗しました。"
        )
        raise SystemExit(1)