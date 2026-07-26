from __future__ import annotations

import argparse
import json
import logging
import os
import tempfile
from pathlib import Path
from typing import Any

from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from notification.notification_calculator import (
    NOTIFICATION_COLUMNS,
    VALID_NOTIFICATION_STATUSES,
)

LOGGER = logging.getLogger(__name__)
LOCK_TIMEOUT_SECONDS = 30
HEADER_FILL_COLOR = "548235"
HEADER_FONT_COLOR = "FFFFFF"
STATUS_FILL_COLORS = {
    "PENDING": "FFF2CC",
    "SUCCESS": "C6E0B4",
    "FAILED": "F4CCCC",
    "SKIPPED": "D9E1F2",
}


class NotificationWriterError(RuntimeError):
    """通知用シートへの書き込みに失敗した場合の例外。"""


def validate_notification_data(
    notification_data: dict[str, list[dict[str, object]]],
) -> None:
    """Calculatorから受け取ったデータ構造と列名を検証する。"""
    if not isinstance(notification_data, dict):
        raise NotificationWriterError(
            "notification_dataはdictである必要があります。"
        )

    missing_sheets = [
        sheet_name
        for sheet_name in NOTIFICATION_COLUMNS
        if sheet_name not in notification_data
    ]
    extra_sheets = [
        sheet_name
        for sheet_name in notification_data
        if sheet_name not in NOTIFICATION_COLUMNS
    ]

    if missing_sheets or extra_sheets:
        raise NotificationWriterError(
            "通知データのシート構成が仕様と一致しません。\n"
            f"不足シート: {missing_sheets}\n"
            f"余分なシート: {extra_sheets}"
        )

    for sheet_name, rows in notification_data.items():
        if not isinstance(rows, list):
            raise NotificationWriterError(
                f"{sheet_name}の値はlistである必要があります。"
            )

        expected_columns = NOTIFICATION_COLUMNS[sheet_name]

        for row_number, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                raise NotificationWriterError(
                    f"{sheet_name}の{row_number}件目はdictである必要があります。"
                )

            missing_columns = [
                column for column in expected_columns if column not in row
            ]
            extra_columns = [
                column for column in row if column not in expected_columns
            ]

            if missing_columns or extra_columns:
                raise NotificationWriterError(
                    f"{sheet_name}の{row_number}件目の項目が仕様と一致しません。\n"
                    f"期待する項目: {expected_columns}\n"
                    f"実際の項目: {list(row.keys())}\n"
                    f"不足項目: {missing_columns}\n"
                    f"余分な項目: {extra_columns}"
                )

            status = str(row.get("notification_status", "")).strip().upper()
            if status not in VALID_NOTIFICATION_STATUSES:
                raise NotificationWriterError(
                    f"{sheet_name}の{row_number}件目に"
                    f"不正なnotification_statusがあります: {status!r}"
                )


def validate_workbook_sheets(workbook: Any) -> None:
    """Excel内に通知用シートがすべて存在することを確認する。"""
    missing_sheets = [
        sheet_name
        for sheet_name in NOTIFICATION_COLUMNS
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise NotificationWriterError(
            "notification_database.xlsxに必要な通知用シートがありません。\n"
            + "\n".join(f"- {sheet_name}" for sheet_name in missing_sheets)
        )


def normalize_excel_value(column_name: str, value: object) -> object:
    """Excelへ書き込む値を列に応じて正規化する。"""
    if value is None:
        if column_name in {
            "comparison_previous_month",
            "monthly_peak_hour",
            "action_effectiveness",
        }:
            return "N/A"
        return None

    if column_name == "notification_status":
        return str(value).strip().upper()

    return value


def replace_notification_sheet_data(
    worksheet: Worksheet,
    columns: list[str],
    rows: list[dict[str, object]],
) -> None:
    """通知用シートをCalculatorの最新結果へ置き換える。"""
    if worksheet.max_row > 0:
        worksheet.delete_rows(idx=1, amount=worksheet.max_row)

    worksheet.append(columns)

    for row in rows:
        worksheet.append([
            normalize_excel_value(column_name=column, value=row.get(column))
            for column in columns
        ])

    format_notification_sheet(worksheet=worksheet, columns=columns)


def format_notification_sheet(
    worksheet: Worksheet,
    columns: list[str],
) -> None:
    """通知用シートの書式を設定する。"""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL_COLOR,
    )
    header_font = Font(
        bold=True,
        color=HEADER_FONT_COLOR,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_positions = {
        column: index + 1 for index, column in enumerate(columns)
    }

    set_number_format(
        worksheet,
        column_positions,
        "confidence",
        "0.0000",
    )

    for column_name in [
        "stay_duration",
        "average_stay_time",
        "familiarity_score",
        "trap_score",
    ]:
        set_number_format(
            worksheet,
            column_positions,
            column_name,
            "0.000",
        )

    for column_name in [
        "boar_ratio",
        "monkey_ratio",
        "comparison_previous_month",
    ]:
        set_number_format(
            worksheet,
            column_positions,
            column_name,
            '0.00"%"',
        )

    for column_name in [
        "total_detection_count",
        "boar_count",
        "monkey_count",
        "monthly_total_count",
        "monthly_peak_hour",
        "year",
        "total_count",
        "peak_hour",
    ]:
        set_number_format(
            worksheet,
            column_positions,
            column_name,
            "0",
        )

    apply_status_validation_and_style(
        worksheet=worksheet,
        column_positions=column_positions,
    )
    adjust_column_widths(worksheet)


def set_number_format(
    worksheet: Worksheet,
    column_positions: dict[str, int],
    column_name: str,
    number_format: str,
) -> None:
    """指定列へExcelの数値表示形式を設定する。"""
    column_index = column_positions.get(column_name)
    if column_index is None:
        return

    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=column_index)
        if isinstance(cell.value, (int, float)):
            cell.number_format = number_format


def apply_status_validation_and_style(
    worksheet: Worksheet,
    column_positions: dict[str, int],
) -> None:
    """notification_status列へ入力制限と色分けを設定する。"""
    status_column_index = column_positions.get("notification_status")
    if status_column_index is None:
        return

    allowed_statuses = sorted(VALID_NOTIFICATION_STATUSES)
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(allowed_statuses) + '"',
        allow_blank=False,
    )
    validation.error = (
        "PENDING、SUCCESS、FAILED、SKIPPEDのいずれかを入力してください。"
    )
    validation.errorTitle = "不正な通知状態"
    validation.prompt = "通知状態を選択してください。"
    validation.promptTitle = "notification_status"

    worksheet.add_data_validation(validation)

    start_cell = worksheet.cell(row=2, column=status_column_index)
    end_cell = worksheet.cell(
        row=max(worksheet.max_row, 1000),
        column=status_column_index,
    )
    validation.add(f"{start_cell.coordinate}:{end_cell.coordinate}")

    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=status_column_index)
        status = str(cell.value or "").strip().upper()
        fill_color = STATUS_FILL_COLORS.get(status)

        if fill_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        cell.alignment = Alignment(horizontal="center", vertical="center")


def adjust_column_widths(worksheet: Worksheet) -> None:
    """セル内容に応じて列幅を調整する。"""
    for column_cells in worksheet.columns:
        maximum_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue
            maximum_length = max(maximum_length, len(str(cell.value)))

        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(
            max(maximum_length + 2, 12),
            40,
        )


def save_workbook_atomically(
    workbook: Any,
    workbook_path: Path,
) -> None:
    """一時ファイルへ保存後、元のExcelと原子的に置き換える。"""
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}_",
        suffix=".xlsx",
        dir=workbook_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)

    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def write_all_notification_sheets(
    workbook_path: Path,
    notification_data: dict[str, list[dict[str, object]]],
) -> None:
    """
    Calculatorの計算結果を5つのnotificationシートへ書き込む。

    rawシートは変更しない。
    """
    workbook_path = Path(workbook_path).resolve()

    if not workbook_path.is_file():
        raise FileNotFoundError(
            "notification_database.xlsxが見つかりません: "
            f"{workbook_path}"
        )

    validate_notification_data(notification_data)

    lock_path = workbook_path.parent / f".{workbook_path.name}.lock"
    lock = FileLock(str(lock_path), timeout=LOCK_TIMEOUT_SECONDS)

    try:
        with lock:
            workbook = load_workbook(
                workbook_path,
                read_only=False,
                data_only=False,
            )

            try:
                validate_workbook_sheets(workbook)

                for sheet_name, columns in NOTIFICATION_COLUMNS.items():
                    rows = notification_data[sheet_name]

                    replace_notification_sheet_data(
                        worksheet=workbook[sheet_name],
                        columns=columns,
                        rows=rows,
                    )

                    LOGGER.info(
                        "%sへ%d件書き込みました。",
                        sheet_name,
                        len(rows),
                    )

                save_workbook_atomically(
                    workbook=workbook,
                    workbook_path=workbook_path,
                )
            finally:
                workbook.close()

    except Timeout as error:
        raise TimeoutError(
            "notification_database.xlsxのロックを"
            f"{LOCK_TIMEOUT_SECONDS}秒以内に取得できませんでした。"
        ) from error

    LOGGER.info(
        "5つのnotificationシートを更新しました: %s",
        workbook_path,
    )


def main() -> None:
    """Calculatorを実行し、その結果をExcelへ書き込む。"""
    from notification.notification_calculator import (
        calculate_all_notifications,
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description=(
            "notification_database.xlsxのrawシートから"
            "通知用データを計算し、notificationシートへ書き込みます。"
        )
    )
    parser.add_argument(
        "workbook_path",
        type=Path,
        help="notification_database.xlsxのパス",
    )
    parser.add_argument(
        "--preview-json",
        action="store_true",
        help="Excel更新前に計算結果をJSON表示します。",
    )

    args = parser.parse_args()

    calculated_data = calculate_all_notifications(
        workbook_path=args.workbook_path,
    )

    if args.preview_json:
        print(
            json.dumps(
                calculated_data,
                ensure_ascii=False,
                indent=2,
                default=str,
            )
        )

    write_all_notification_sheets(
        workbook_path=args.workbook_path,
        notification_data=calculated_data,
    )


if __name__ == "__main__":
    try:
        main()
    except PermissionError:
        LOGGER.error(
            "notification_database.xlsxへ書き込めません。"
            "Excelで開いている場合は閉じてください。"
        )
        raise SystemExit(1)
    except (
        FileNotFoundError,
        NotificationWriterError,
        TimeoutError,
        ValueError,
    ) as error:
        LOGGER.error("%s", error)
        raise SystemExit(1)
    except KeyboardInterrupt:
        LOGGER.info("Writerを終了しました。")
        raise SystemExit(0)
    except Exception:
        LOGGER.exception("notificationシートの更新に失敗しました。")
        raise SystemExit(1)