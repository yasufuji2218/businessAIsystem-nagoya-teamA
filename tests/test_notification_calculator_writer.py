from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from notification.notification_calculator import calculate_all_notifications
from notification.notification_writer import write_all_notification_sheets


def count_data_rows(
    worksheet: Worksheet,
) -> int:
    """
    1列目に値がある行だけを実データとして数える。

    Excelの入力規則や書式設定によってmax_rowが増えても、
    空行をテスト対象に含めない。
    """
    return sum(
        1
        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[0] is not None
    )


def test_calculator_creates_expected_data(
    sample_workbook_path,
):
    result = calculate_all_notifications(
        sample_workbook_path
    )

    assert len(result["realtime_notification"]) == 3
    assert len(result["daily_notification"]) == 1
    assert len(result["weekly_notification"]) == 2
    assert len(result["monthly_notification"]) == 2
    assert len(result["yearly_notification"]) == 1

    daily = result["daily_notification"][0]

    assert daily["date"] == "2026-07-21"
    assert daily["total_detection_count"] == 11
    assert daily["boar_count"] == 1
    assert daily["monkey_count"] == 10
    assert daily["average_stay_time"] == 0.457
    assert daily["night_alert_level"] == "HIGH"
    assert daily["notification_status"] == "PENDING"

    july = next(
        row
        for row in result["monthly_notification"]
        if row["month"] == "2026-07"
    )

    assert july["monthly_total_count"] == 11
    assert july["boar_ratio"] == 9.09
    assert july["monkey_ratio"] == 90.91
    assert july["comparison_previous_month"] == 120.0
    assert july["monthly_peak_hour"] == 10
    assert july["action_effectiveness"] is None
    assert july["trap_score"] == 0.8
    assert july["notification_status"] == "PENDING"


def test_writer_writes_notification_sheets(
    sample_workbook_path,
):
    calculated_data = calculate_all_notifications(
        sample_workbook_path
    )

    write_all_notification_sheets(
        workbook_path=sample_workbook_path,
        notification_data=calculated_data,
    )

    workbook = load_workbook(
        sample_workbook_path,
        data_only=False,
    )

    try:
        assert count_data_rows(
            workbook["realtime_notification"]
        ) == 3

        assert count_data_rows(
            workbook["daily_notification"]
        ) == 1

        assert count_data_rows(
            workbook["weekly_notification"]
        ) == 2

        assert count_data_rows(
            workbook["monthly_notification"]
        ) == 2

        assert count_data_rows(
            workbook["yearly_notification"]
        ) == 1

        daily = workbook["daily_notification"]

        assert daily["A2"].value == "2026-07-21"
        assert daily["B2"].value == 11
        assert daily["C2"].value == 1
        assert daily["D2"].value == 10
        assert daily["E2"].value == 0.457
        assert daily["F2"].value == "HIGH"
        assert daily["G2"].value == "PENDING"

        monthly = workbook["monthly_notification"]

        # 2026-06が2行目、2026-07が3行目
        assert monthly["A3"].value == "2026-07"
        assert monthly["B3"].value == "CAM001"
        assert monthly["C3"].value == 11
        assert monthly["D3"].value == 9.09
        assert monthly["E3"].value == 90.91
        assert monthly["F3"].value == 120.0
        assert monthly["G3"].value == 10
        assert monthly["H3"].value == "N/A"
        assert monthly["I3"].value == 0.8
        assert monthly["J3"].value == "PENDING"

        # 修正版Writerでは空の1000行目を実セルとして作らない。
        assert workbook["realtime_notification"].max_row == 4
        assert workbook["daily_notification"].max_row == 2
        assert workbook["weekly_notification"].max_row == 3
        assert workbook["monthly_notification"].max_row == 3
        assert workbook["yearly_notification"].max_row == 2

    finally:
        workbook.close()
