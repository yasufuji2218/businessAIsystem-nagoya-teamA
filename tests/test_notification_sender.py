from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from notification.notification_calculator import calculate_all_notifications
from notification.notification_sender import (
    SlackSendResult,
    send_pending_notifications,
)
from notification.notification_writer import write_all_notification_sheets


NOTIFICATION_SHEET_NAMES = (
    "realtime_notification",
    "daily_notification",
    "weekly_notification",
    "monthly_notification",
    "yearly_notification",
)


def get_data_row_numbers(
    worksheet: Worksheet,
) -> list[int]:
    """
    1列目に値がある実データ行の行番号だけを返す。

    書式や入力規則だけが設定された空行は除外する。
    """
    return [
        row_number
        for row_number in range(
            2,
            worksheet.max_row + 1,
        )
        if worksheet.cell(
            row=row_number,
            column=1,
        ).value is not None
    ]


def test_sender_updates_status_without_real_slack(
    sample_workbook_path,
    monkeypatch,
):
    calculated_data = calculate_all_notifications(
        sample_workbook_path
    )

    write_all_notification_sheets(
        workbook_path=sample_workbook_path,
        notification_data=calculated_data,
    )

    monkeypatch.setenv(
        "SLACK_WEBHOOK_URL",
        "https://hooks.slack.com/services/TEST/TEST/TEST",
    )

    monkeypatch.setattr(
        "notification.notification_sender.load_environment",
        lambda: None,
    )

    sent_messages: list[str] = []

    def fake_send_with_retry(
        webhook_url: str,
        message: str,
    ) -> SlackSendResult:
        sent_messages.append(message)

        return SlackSendResult(
            success=True,
            status_code=200,
            response_text="ok",
            error_message=None,
        )

    monkeypatch.setattr(
        "notification.notification_sender.send_with_retry",
        fake_send_with_retry,
    )

    result = send_pending_notifications(
        sample_workbook_path
    )

    assert result["success"] == 9
    assert result["failed"] == 0
    assert result["total"] == 9
    assert len(sent_messages) == 9

    workbook = load_workbook(
        sample_workbook_path,
        data_only=False,
    )

    try:
        checked_rows = 0

        for sheet_name in NOTIFICATION_SHEET_NAMES:
            worksheet = workbook[sheet_name]
            status_column = worksheet.max_column
            data_row_numbers = get_data_row_numbers(
                worksheet
            )

            for row_number in data_row_numbers:
                assert worksheet.cell(
                    row=row_number,
                    column=status_column,
                ).value == "SUCCESS"

                checked_rows += 1

        assert checked_rows == 9

    finally:
        workbook.close()


def test_sender_does_not_resend_success_rows(
    sample_workbook_path,
    monkeypatch,
):
    calculated_data = calculate_all_notifications(
        sample_workbook_path
    )

    write_all_notification_sheets(
        workbook_path=sample_workbook_path,
        notification_data=calculated_data,
    )

    monkeypatch.setenv(
        "SLACK_WEBHOOK_URL",
        "https://hooks.slack.com/services/TEST/TEST/TEST",
    )

    monkeypatch.setattr(
        "notification.notification_sender.load_environment",
        lambda: None,
    )

    send_count = 0

    def fake_send_with_retry(
        webhook_url: str,
        message: str,
    ) -> SlackSendResult:
        nonlocal send_count
        send_count += 1

        return SlackSendResult(
            success=True,
            status_code=200,
            response_text="ok",
            error_message=None,
        )

    monkeypatch.setattr(
        "notification.notification_sender.send_with_retry",
        fake_send_with_retry,
    )

    first_result = send_pending_notifications(
        sample_workbook_path
    )

    second_result = send_pending_notifications(
        sample_workbook_path
    )

    assert first_result["success"] == 9
    assert first_result["total"] == 9

    assert second_result["success"] == 0
    assert second_result["failed"] == 0
    assert second_result["total"] == 0
    assert second_result["skipped"] == 9

    # 2回目はSlack送信関数が呼ばれていない。
    assert send_count == 9
