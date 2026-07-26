from openpyxl import load_workbook

import notification.notification_database_updater as updater


def test_updater_replaces_raw_sheets(
    sample_workbook_path,
    backend_csv_dir,
    monkeypatch,
):
    result_dir = sample_workbook_path.parent

    monkeypatch.setattr(updater, "RESULT_DIR", result_dir)
    monkeypatch.setattr(updater, "DATABASE_XLSX", sample_workbook_path)
    monkeypatch.setattr(
        updater,
        "DATABASE_LOCK_FILE",
        result_dir / ".notification_database.xlsx.lock",
    )
    monkeypatch.setattr(updater, "BACKEND_DIR", backend_csv_dir)

    for filename, setting in updater.SOURCE_SETTINGS.items():
        setting["path"] = backend_csv_dir / filename

    result = updater.update_notification_database()
    assert result == sample_workbook_path.resolve()

    workbook = load_workbook(sample_workbook_path)
    try:
        assert workbook["realtime_sheet"].max_row == 2
        assert workbook["daily_sheet"].max_row == 2
        assert workbook["weekly_sheet"].max_row == 2
        assert workbook["monthly_sheet"].max_row == 2
        assert workbook["yearly_sheet"].max_row == 2
        assert workbook["daily_notification"].max_row == 1
    finally:
        workbook.close()
